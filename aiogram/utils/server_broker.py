import logging
from datetime import datetime
from io import BytesIO
from json import dumps
from typing import Union

import aiohttp
from constants import SECRET, SERVER_URL
from openpyxl import Workbook


class ServerBroker:
    def __init__(self):
        self.url = SERVER_URL
        self.headers = {"secret": SECRET, "Content-Type": "application/json"}

    async def user_identical(self, user_id: int, name: str) -> bool:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.get(f"{self.url}/users/{user_id}")

            response.raise_for_status()

            json = await response.json()

            if "status" in json and not json["status"]:
                return False

            if "status" in json and json["name"] != name:
                return False

            return True

    async def add_user(self, user_id: int, name: str) -> bool:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            if not await self.user_identical(user_id, name):
                response = await ses.post(
                    f"{self.url}/users/add", json={"id": user_id, "name": name}
                )

                response.raise_for_status()

                json = await response.json()

                if "status" in json and not json["status"]:
                    return False

                return True

        return True

    async def add_expenses(
        self, user_id: int, name: str, date: datetime, amount: int
    ) -> str:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.post(
                f"{self.url}/expenses/add",
                json={
                    "name": name,
                    "amount_uah": amount,
                    "user_id": user_id,
                    "created_at": date.strftime("%Y-%m-%d"),
                },
            )

            response.raise_for_status()

            json = await response.json()

            if "status" in json and not json["status"]:
                logging.error(json)
            else:
                json = json["message"]
                return f"\nНазва: {json['name']}\nСума в гривнях: {json['amount_uah']}₴\nСума в долларах: {round(json['amount_usd'], 2)}$\nДата: {datetime.strptime(json['created_at'], '%Y-%m-%d').strftime('%d.%m.%Y')}"

    async def _report(self, data: list[dict], with_id: bool = False) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Звіт"

        if with_id:
            ws.append(["Назва", "Сума в гривнях", "Сума в долларах", "Дата", "ID"])
            ws.column_dimensions["E"].width = 45
        else:
            ws.append(["Назва", "Сума в гривнях", "Сума в долларах", "Дата"])

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 20

        for row in data:
            if with_id:
                ws.append(list(row.values()))
            else:
                ws.append(list(row.values())[:-1])

        total_row = len(data) + 2

        ws[f"B{total_row}"] = "=SUM(B2:B{})".format(total_row - 1)
        ws[f"C{total_row}"] = "=SUM(C2:C{})".format(total_row - 1)

        output = BytesIO()

        wb.save(output)
        output.seek(0)

        return output

    async def generate_report(self, user_id: int) -> BytesIO:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.get(f"{self.url}/expenses/{user_id}")

            response.raise_for_status()

            json = await response.json()

            return await self._report(json, True)

    async def generate_report_from_range(
        self, user_id: int, from_range: datetime, to_range: datetime
    ) -> BytesIO:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.get(
                f"{self.url}/expenses/{user_id}/{from_range.strftime('%d.%m.%Y')}/{to_range.strftime('%d.%m.%Y')}"
            )

            response.raise_for_status()

            json = await response.json()

            return await self._report(json)

    async def delete(self, id: str) -> bool:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.delete(f"{self.url}/expenses/{id}")

            response.raise_for_status()

            json = await response.json()

            if "status" in json and json["status"]:
                return True

            return False

    async def delete_many(self, ids: list[str]) -> bool:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.delete(f"{self.url}/expenses", data=dumps(ids))

            response.raise_for_status()

            json = await response.json()

            if "status" in json and json["status"]:
                return True

            return False

    async def has_expenses(self, id: str) -> Union[bool, dict]:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.get(f"{self.url}/expenses/get/{id}")

            response.raise_for_status()

            json = await response.json()

            if "status" in json and not json["status"]:
                return False
            else:
                return json

    async def update_expenses(
        self, id: str, name: str, amount: str
    ) -> Union[bool, dict]:
        async with aiohttp.ClientSession(headers=self.headers) as ses:
            response = await ses.patch(
                f"{self.url}/expenses/{id}",
                data=dumps({"name": name, "amount_uah": amount}),
            )

            response.raise_for_status()

            json = await response.json()

            if "status" in json and not json["status"]:
                return False
            else:
                return json
